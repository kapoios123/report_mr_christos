"""
Dynamic Excel Formatter Pipeline
app.py — Core Streamlit Application
"""

import io
import json
import logging
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent
SETTINGS_PATH  = BASE_DIR / "settings.json"
GITHUB_RAW_URL = "https://raw.githubusercontent.com/kapoios123/report_mr_christos/main/rules.json"
SYNC_TIMEOUT   = 3
LOG_FILE       = BASE_DIR / "formatter.log"

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)

# Στήλες που η ΛΟΓΙΚΗ χρειάζεται — δεν μπορούν να αφαιρεθούν από τον χρήστη
CRITICAL_COLUMNS = [
    "ΠΡΟΜΗΘΕΥΤΗΣ",
    "Container",
    "Ναυτιλιακή",
    "Ημ. Αναμεν.Παράδ",
]

# Default στήλες που κρατάμε (αποθηκεύονται στο settings.json)
DEFAULT_COLUMNS_TO_KEEP = [
    "ΑΠΟΘ.",
    "ΠΕΡΙΓΡΑΦΗ ΕΙΔΟΥΣ",
    "Διάσταση",
    "ΠΟΣ. ΕΚΚΡΕΜΗΣ",
    "TIMH",
    "ΠΡΟΜΗΘΕΥΤΗΣ",
    "Container",
    "Ναυτιλιακή",
    "Ημ/νία Φόρτωσης",
    "Ημ. Αναμεν.Παράδ",
]

# Ονόματα στηλών ΜΕΤΑ τη διαγραφή
COL_SUPPLIER   = "ΠΡΟΜΗΘΕΥΤΗΣ"
COL_DELIVERY   = "Ημ. Αναμεν.Παράδ"
COL_CONTAINER  = "Container"
COL_SHIPPING   = "Ναυτιλιακή"

SHEET_EUROPE   = "Europe"
SHEET_CHINA    = "China"

# ─────────────────────────────────────────────
# SETTINGS (αποθηκεύονται τοπικά)
# ─────────────────────────────────────────────
DEFAULT_SETTINGS = {
    "supplier_mappings": [],
    "columns_to_keep": DEFAULT_COLUMNS_TO_KEEP,
}

def load_settings() -> dict:
    if SETTINGS_PATH.exists():
        try:
            with open(SETTINGS_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return DEFAULT_SETTINGS.copy()

def save_settings(settings: dict):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2, ensure_ascii=False)

# ─────────────────────────────────────────────
# GITHUB SYNC (για μελλοντικές αναβαθμίσεις)
# ─────────────────────────────────────────────
def sync_from_github() -> str:
    try:
        resp = requests.get(GITHUB_RAW_URL, timeout=SYNC_TIMEOUT)
        resp.raise_for_status()
        log.info("GitHub sync OK.")
        return "✅ Σύνδεση με GitHub επιτυχής."
    except Exception as e:
        log.warning(f"GitHub sync failed: {e}")
        return "⚠️ Offline mode — χρήση τοπικών ρυθμίσεων."

# ─────────────────────────────────────────────
# EXCEL PROCESSING
# ─────────────────────────────────────────────
def process_excel(raw_bytes: bytes, settings: dict) -> bytes:
    mappings = {
        m["original"]: {"replacement": m["replacement"], "region": m["region"]}
        for m in settings.get("supplier_mappings", [])
    }

    # ── 1. Φόρτωση με pandas — διαβάζουμε από γραμμή 2 (header=1, 0-indexed) ──
    df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=1, dtype=str)

    # ── 2. Καθαρισμός ονομάτων επικεφαλίδων ──
    df_raw.columns = [str(c).strip() for c in df_raw.columns]

    # ── 3. Κρατάμε ΜΟΝΟ τις στήλες που έχει ορίσει ο χρήστης ──
    columns_to_keep = settings.get("columns_to_keep", DEFAULT_COLUMNS_TO_KEEP)
    keep    = [c for c in columns_to_keep if c in df_raw.columns]
    missing = [c for c in columns_to_keep if c not in df_raw.columns]
    critical_missing = [c for c in CRITICAL_COLUMNS if c not in df_raw.columns]

    if critical_missing:
        raise ValueError(
            f"⛔ Κρίσιμες στήλες δεν βρέθηκαν στο αρχείο: {', '.join(critical_missing)}\n"
            f"Έλεγξε ότι το αρχείο είναι σωστό ή ενημέρωσε τα ονόματα στήλων."
        )
    if missing:
        log.warning(f"Στήλες που δεν βρέθηκαν (αγνοούνται): {missing}")

    df = df_raw[keep].copy()

    # ── 5. Αντικατάσταση προμηθευτών & προσθήκη στήλης Region ──
    if COL_SUPPLIER not in df.columns:
        raise ValueError(f"Δεν βρέθηκε η στήλη '{COL_SUPPLIER}'. "
                         f"Διαθέσιμες: {list(df.columns)}")

    def map_supplier(val):
        if pd.isna(val):
            return val, "Unknown"
        val_str = str(val).strip()
        if val_str in mappings:
            m = mappings[val_str]
            return m["replacement"], m["region"]
        return val_str, "Unknown"

    df[[COL_SUPPLIER, "Region"]] = df[COL_SUPPLIER].apply(
        lambda v: pd.Series(map_supplier(v))
    )

    # ── 6. Διαχωρισμός σε Europe / China ──
    df_europe = df[df["Region"] == "Europe"].copy()
    df_china  = df[df["Region"] == "China"].copy()

    # ── 7. Sort function για κάθε sheet ──
    def sort_df(frame: pd.DataFrame) -> pd.DataFrame:
        if frame.empty:
            return frame

        has_delivery = COL_DELIVERY in frame.columns
        has_container = COL_CONTAINER in frame.columns
        has_shipping  = COL_SHIPPING  in frame.columns

        # Μετατροπή ημερομηνίας
        if has_delivery:
            frame["_delivery_dt"] = pd.to_datetime(
                frame[COL_DELIVERY], errors="coerce", dayfirst=True
            )

        def sort_priority(row):
            has_cont = has_container and pd.notna(row.get(COL_CONTAINER)) and str(row.get(COL_CONTAINER, "")).strip() != ""
            has_ship = has_shipping  and pd.notna(row.get(COL_SHIPPING))  and str(row.get(COL_SHIPPING,  "")).strip() != ""
            has_date = has_delivery  and pd.notna(row.get(COL_DELIVERY))

            if (has_cont or has_ship) and has_date:
                return 0   # Πρώτες: container/ναυτιλιακή + ημερομηνία
            elif has_date:
                return 1   # Δεύτερες: μόνο ημερομηνία
            else:
                return 2   # Τελευταίες: χωρίς ημερομηνία

        frame["_sort_priority"] = frame.apply(sort_priority, axis=1)

        if has_delivery:
            frame = frame.sort_values(
                by=["_sort_priority", "_delivery_dt"],
                ascending=[True, True],
                na_position="last"
            )
            frame = frame.drop(columns=["_delivery_dt"])
        else:
            frame = frame.sort_values(by=["_sort_priority"], ascending=True)

        frame = frame.drop(columns=["_sort_priority"], errors="ignore")
        return frame

    df_europe = sort_df(df_europe)
    df_china  = sort_df(df_china)

    # Αφαίρεση βοηθητικής στήλης Region
    df_europe = df_europe.drop(columns=["Region"], errors="ignore")
    df_china  = df_china.drop(columns=["Region"], errors="ignore")

    # ── 8. Εγγραφή στο Excel με 2 sheets ──
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_europe.to_excel(writer, sheet_name=SHEET_EUROPE, index=False)
        df_china.to_excel(writer, sheet_name=SHEET_CHINA,  index=False)

        # ── 9. Βασικό formatting και στα 2 sheets ──
        for sheet_name in [SHEET_EUROPE, SHEET_CHINA]:
            ws_out = writer.sheets[sheet_name]
            apply_basic_formatting(ws_out)

    out.seek(0)
    return out.read()


def apply_basic_formatting(ws):
    """Εφαρμόζει header styling, column widths, freeze και auto-filter."""
    header_fill = PatternFill("solid", fgColor="FF2E4057")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    cell_font   = Font(name="Calibri", size=10)
    border_side = Side(style="thin", color="FFBFBFBF")
    border      = Border(left=border_side, right=border_side,
                         top=border_side, bottom=border_side)
    alt_fill    = PatternFill("solid", fgColor="FFEDF2F7")
    base_fill   = PatternFill("solid", fgColor="FFFFFFFF")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        # Auto column width
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = base_fill if row_idx % 2 == 0 else alt_fill
        for cell in row:
            cell.fill      = fill
            cell.font      = cell_font
            cell.border    = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Excel Report Pipeline",
        page_icon="📊",
        layout="wide",
    )

    st.markdown("""
    <style>
        .stButton>button {
            background-color: #2E4057; color: white;
            border-radius: 6px; padding: 0.5em 1.5em;
            font-weight: 600; border: none;
        }
        .stButton>button:hover { background-color: #3A5270; }
        .mapping-card {
            background: white; border-left: 4px solid #2E4057;
            padding: 8px 14px; border-radius: 4px;
            margin-bottom: 6px; font-size: 0.88em;
        }
    </style>
    """, unsafe_allow_html=True)

    st.title("📊 Excel Report Pipeline")
    st.caption("Καθαρισμός, μετασχηματισμός και διαχωρισμός Excel report σε Europe / China.")

    # ── Load settings ──
    if "settings" not in st.session_state:
        st.session_state.settings = load_settings()

    settings = st.session_state.settings

    # ════════════════════════════════════════════
    # SIDEBAR — Ρυθμίσεις Προμηθευτών
    # ════════════════════════════════════════════
    with st.sidebar:
        st.header("⚙️ Ρυθμίσεις Προμηθευτών")
        st.caption("Οι ρυθμίσεις αποθηκεύονται τοπικά και παραμένουν για πάντα.")

        # ── GitHub sync status ──
        if "synced" not in st.session_state:
            st.session_state.synced = sync_from_github()
        st.info(st.session_state.synced)

        st.divider()

        # ── Υπάρχουσες αντιστοιχίσεις ──
        mappings = settings.get("supplier_mappings", [])

        if mappings:
            st.subheader(f"📋 Αντιστοιχίσεις ({len(mappings)})")
            for i, m in enumerate(mappings):
                region_icon = "🇪🇺" if m["region"] == "Europe" else "🇨🇳"
                col_a, col_b = st.columns([4, 1])
                with col_a:
                    st.markdown(f"""
                    <div class='mapping-card'>
                    {region_icon} <b>{m['original']}</b><br>
                    → <span style='color:#2E4057'>{m['replacement']}</span>
                    </div>
                    """, unsafe_allow_html=True)
                with col_b:
                    if st.button("🗑️", key=f"del_{i}", help="Διαγραφή"):
                        settings["supplier_mappings"].pop(i)
                        save_settings(settings)
                        st.rerun()
        else:
            st.info("Δεν υπάρχουν αντιστοιχίσεις ακόμα.")

        st.divider()

        # ── Διαχείριση Στηλών ──
        st.subheader("📐 Στήλες Output")
        st.caption("Οι 🔒 στήλες είναι κρίσιμες και δεν μπορούν να αφαιρεθούν.")

        current_cols = settings.get("columns_to_keep", DEFAULT_COLUMNS_TO_KEEP)

        # Εμφάνιση υπαρχουσών στηλών με κουμπί διαγραφής
        for i, col in enumerate(current_cols):
            is_critical = col in CRITICAL_COLUMNS
            c1, c2 = st.columns([5, 1])
            with c1:
                icon = "🔒" if is_critical else "📋"
                st.markdown(f"{icon} `{col}`")
            with c2:
                if is_critical:
                    st.markdown("—")
                else:
                    if st.button("🗑️", key=f"delcol_{i}", help="Αφαίρεση στήλης"):
                        settings["columns_to_keep"].remove(col)
                        save_settings(settings)
                        st.rerun()

        st.markdown("")
        new_col = st.text_input("Προσθήκη στήλης", placeholder="Ακριβές όνομα επικεφαλίδας",
                                 key="new_col_input")
        if st.button("➕ Προσθήκη", width="stretch", key="add_col_btn"):
            new_col = new_col.strip()
            if not new_col:
                st.error("Συμπλήρωσε όνομα στήλης.")
            elif new_col in current_cols:
                st.warning("Η στήλη υπάρχει ήδη.")
            else:
                settings["columns_to_keep"].append(new_col)
                save_settings(settings)
                st.success(f"✅ Προστέθηκε: {new_col}")
                st.rerun()

        st.divider()
        st.subheader("➕ Νέα Αντιστοίχιση")
        new_original    = st.text_input("Αρχικό όνομα προμηθευτή", key="new_orig",
                                         placeholder="π.χ. XINYI GROUP(GLASS) COMPANY LIMITED")
        new_replacement = st.text_input("Αντικατάσταση", key="new_repl",
                                         placeholder="π.χ. XINYI")
        new_region      = st.radio("Περιοχή", ["Europe", "China"],
                                    horizontal=True, key="new_region")

        if st.button("💾 Αποθήκευση Αντιστοίχισης", width="stretch"):
            if new_original.strip() and new_replacement.strip():
                # Έλεγχος αν υπάρχει ήδη
                existing = [m["original"] for m in mappings]
                if new_original.strip() in existing:
                    st.warning("Αυτός ο προμηθευτής υπάρχει ήδη. Διέγραψέ τον πρώτα για να τον επεξεργαστείς.")
                else:
                    settings["supplier_mappings"].append({
                        "original"   : new_original.strip(),
                        "replacement": new_replacement.strip(),
                        "region"     : new_region,
                    })
                    save_settings(settings)
                    st.success(f"✅ Αποθηκεύτηκε: {new_original.strip()} → {new_replacement.strip()}")
                    st.rerun()
            else:
                st.error("Συμπλήρωσε και τα δύο πεδία.")

    # ════════════════════════════════════════════
    # MAIN — Upload & Process
    # ════════════════════════════════════════════
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("📁 Ανέβασε το Raw Excel")
        uploaded = st.file_uploader(
            "Επίλεξε αρχείο .xlsx",
            type=["xlsx"],
        )

    with col2:
        st.subheader("ℹ️ Τι γίνεται αυτόματα")
        st.markdown("""
        1. Διαγραφή γραμμής 1
        2. Διαγραφή στηλών B D G H J K L M O P U+
        3. Αντικατάσταση προμηθευτών
        4. Διαχωρισμός σε **Europe** / **China**
        5. Sort βάσει ημ. παράδοσης
        6. Μορφοποίηση & download
        """)

    if uploaded:
        raw_bytes = uploaded.read()

        # Preview
        try:
            df_prev = pd.read_excel(io.BytesIO(raw_bytes), nrows=6, header=1)
            st.subheader("🔍 Preview (πρώτες 6 γραμμές, raw)")
            st.dataframe(df_prev, use_container_width=True)
        except Exception as e:
            st.warning(f"Preview error: {e}")

        st.divider()

        # Έλεγχος στηλών πριν την επεξεργασία
        try:
            df_check = pd.read_excel(io.BytesIO(raw_bytes), header=1, nrows=0, dtype=str)
            df_check.columns = [str(c).strip() for c in df_check.columns]
            current_cols = settings.get("columns_to_keep", DEFAULT_COLUMNS_TO_KEEP)
            missing_cols = [c for c in current_cols if c not in df_check.columns]
            critical_missing = [c for c in CRITICAL_COLUMNS if c not in df_check.columns]

            if critical_missing:
                st.error(f"⛔ **Κρίσιμες στήλες λείπουν από το αρχείο:** {', '.join(critical_missing)}\n\nΔεν μπορεί να γίνει επεξεργασία.")
            elif missing_cols:
                st.warning(f"⚠️ Οι παρακάτω στήλες δεν βρέθηκαν και θα αγνοηθούν: `{'`, `'.join(missing_cols)}`")
        except Exception:
            pass

            st.warning("⚠️ Δεν έχεις ορίσει αντιστοιχίσεις προμηθευτών. Προσέθεσε τουλάχιστον μία από το sidebar.")

        if st.button("⚙️ Επεξεργασία & Δημιουργία Report", width="stretch"):
            with st.spinner("Επεξεργασία…"):
                try:
                    result = process_excel(raw_bytes, settings)
                    st.session_state["result"]   = result
                    st.session_state["out_name"] = uploaded.name.replace(".xlsx", "_report.xlsx")
                    st.success("✅ Έτοιμο! Κατέβασε το αρχείο παρακάτω.")
                except Exception as e:
                    log.error(f"Processing error: {e}", exc_info=True)
                    st.error(f"❌ Σφάλμα: {e}")

        if "result" in st.session_state:
            st.download_button(
                label="⬇️ Κατέβασε το Report (.xlsx)",
                data=st.session_state["result"],
                file_name=st.session_state["out_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
    else:
        st.info("👆 Ανέβασε ένα αρχείο Excel για να ξεκινήσεις.")

    st.divider()
    st.caption(f"Excel Report Pipeline · {datetime.now().strftime('%Y-%m-%d')}")


if __name__ == "__main__":
    main()